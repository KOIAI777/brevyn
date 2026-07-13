#!/usr/bin/env python3
"""Audit an XLSX workbook for Brevyn Office QA."""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


EXCEL_ERRORS = ["#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A"]


def main() -> int:
    parser = argparse.ArgumentParser(description="Audit an XLSX workbook and emit JSON.")
    parser.add_argument("xlsx", help="Input .xlsx/.xlsm file")
    parser.add_argument("--json", action="store_true", help="Print JSON report only")
    args = parser.parse_args()

    path = Path(args.xlsx).expanduser().resolve()
    if not path.exists() or path.suffix.lower() not in {".xlsx", ".xlsm"}:
        return fail(f"Invalid XLSX/XLSM file: {path}", args.json)

    report = audit_workbook(path)
    if args.json:
        print(json.dumps(report, indent=2, ensure_ascii=False))
    else:
        print(f"XLSX audit: {report['status']}")
        print(json.dumps(report, indent=2, ensure_ascii=False))
    return 0 if report["ok"] else 1


def audit_workbook(path: Path) -> dict[str, Any]:
    try:
        formulas_wb = load_workbook(path, data_only=False, read_only=False, keep_vba=path.suffix.lower() == ".xlsm")
        values_wb = load_workbook(path, data_only=True, read_only=False, keep_vba=path.suffix.lower() == ".xlsm")
    except Exception as error:  # noqa: BLE001
        return {
            "ok": False,
            "status": "load_failed",
            "input": str(path),
            "warnings": [str(error)],
        }

    try:
        sheets = []
        total_formulas = 0
        formula_errors: dict[str, list[str]] = {error: [] for error in EXCEL_ERRORS}
        blank_sheets: list[str] = []
        hidden_sheets: list[str] = []
        charts_total = 0
        tables_total = 0

        for sheet_name in formulas_wb.sheetnames:
            ws = formulas_wb[sheet_name]
            values_ws = values_wb[sheet_name]
            used_range = worksheet_used_range(ws)
            formula_count = 0
            error_count = 0

            if ws.sheet_state != "visible":
                hidden_sheets.append(sheet_name)
            if used_range == "":
                blank_sheets.append(sheet_name)

            for row in ws.iter_rows():
                for cell in row:
                    value = cell.value
                    if isinstance(value, str) and value.startswith("="):
                        formula_count += 1
                    value_cell = values_ws[cell.coordinate]
                    if isinstance(value_cell.value, str):
                        for error in EXCEL_ERRORS:
                            if error in value_cell.value:
                                formula_errors[error].append(f"{sheet_name}!{cell.coordinate}")
                                error_count += 1
                                break

            tables = list(getattr(ws, "tables", {}).keys())
            charts = getattr(ws, "_charts", []) or []
            total_formulas += formula_count
            charts_total += len(charts)
            tables_total += len(tables)
            sheets.append({
                "name": sheet_name,
                "state": ws.sheet_state,
                "usedRange": used_range,
                "maxRow": ws.max_row,
                "maxColumn": ws.max_column,
                "formulaCount": formula_count,
                "formulaErrorCount": error_count,
                "tableCount": len(tables),
                "tables": tables[:20],
                "chartCount": len(charts),
            })

        compact_errors = {
            error: {"count": len(locations), "locations": locations[:50]}
            for error, locations in formula_errors.items()
            if locations
        }
        warnings = []
        if compact_errors:
            warnings.append("Formula errors found")
        if blank_sheets:
            warnings.append(f"{len(blank_sheets)} blank sheet(s) found")

        ok = not compact_errors
        return {
            "ok": ok,
            "status": "ok" if ok and not warnings else "warnings" if ok else "errors_found",
            "input": str(path),
            "sheetCount": len(sheets),
            "sheets": sheets,
            "hiddenSheets": hidden_sheets,
            "blankSheets": blank_sheets,
            "totalFormulas": total_formulas,
            "totalTables": tables_total,
            "totalCharts": charts_total,
            "formulaErrors": compact_errors,
            "warnings": warnings,
        }
    finally:
        formulas_wb.close()
        values_wb.close()


def worksheet_used_range(ws: Any) -> str:
    cells = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                cells.append(cell)
    if not cells:
        return ""
    min_row = min(cell.row for cell in cells)
    max_row = max(cell.row for cell in cells)
    min_col = min(cell.column for cell in cells)
    max_col = max(cell.column for cell in cells)
    return f"{ws.cell(min_row, min_col).coordinate}:{ws.cell(max_row, max_col).coordinate}"


def fail(message: str, json_only: bool) -> int:
    payload = {"ok": False, "status": "invalid_input", "warnings": [message]}
    print(json.dumps(payload, indent=2) if json_only else message, file=sys.stderr)
    return 1


if __name__ == "__main__":
    raise SystemExit(main())
