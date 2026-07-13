"""Recalculate Excel formulas with LibreOffice and report formula errors."""

import json
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path

from office.soffice import get_soffice_env, resolve_soffice
from openpyxl import load_workbook

EXCEL_ERRORS = ["#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A"]


def recalc(filename, timeout=30):
    input_path = Path(filename).expanduser().resolve()
    if not input_path.exists():
        return {"error": f"File {filename} does not exist"}
    if input_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        return {"error": "Formula recalculation supports .xlsx and .xlsm files"}

    soffice = resolve_soffice()
    if not soffice:
        return {"error": "LibreOffice runtime is unavailable"}

    output_format = "xlsm" if input_path.suffix.lower() == ".xlsm" else "xlsx"
    with tempfile.TemporaryDirectory(prefix="brevyn-xlsx-recalc-") as temp_value:
        temp_dir = Path(temp_value)
        output_dir = temp_dir / "output"
        profile_dir = temp_dir / "profile"
        output_dir.mkdir()
        profile_dir.mkdir()
        command = [
            soffice,
            "--headless",
            "--norestore",
            "--nofirststartwizard",
            f"-env:UserInstallation={profile_dir.resolve().as_uri()}",
            "--convert-to",
            output_format,
            "--outdir",
            str(output_dir),
            str(input_path),
        ]
        try:
            conversion = subprocess.run(
                command,
                capture_output=True,
                text=True,
                env=get_soffice_env(),
                timeout=timeout,
                check=False,
            )
        except subprocess.TimeoutExpired:
            return {"error": f"LibreOffice recalculation timed out after {timeout} seconds"}

        recalculated = output_dir / input_path.name
        if conversion.returncode != 0 or not recalculated.exists():
            message = (
                conversion.stderr
                or conversion.stdout
                or "LibreOffice did not produce a recalculated workbook"
            ).strip()
            return {"error": message}
        shutil.copy2(recalculated, input_path)

    try:
        values_workbook = load_workbook(input_path, data_only=True, keep_vba=input_path.suffix.lower() == ".xlsm")
        error_details = {error: [] for error in EXCEL_ERRORS}
        total_errors = 0
        for sheet in values_workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if not isinstance(cell.value, str):
                        continue
                    for error in EXCEL_ERRORS:
                        if error in cell.value:
                            error_details[error].append(f"{sheet.title}!{cell.coordinate}")
                            total_errors += 1
                            break
        values_workbook.close()

        formula_workbook = load_workbook(input_path, data_only=False, keep_vba=input_path.suffix.lower() == ".xlsm")
        total_formulas = sum(
            1
            for sheet in formula_workbook.worksheets
            for row in sheet.iter_rows()
            for cell in row
            if isinstance(cell.value, str) and cell.value.startswith("=")
        )
        formula_workbook.close()

        summary = {
            error: {"count": len(locations), "locations": locations[:20]}
            for error, locations in error_details.items()
            if locations
        }
        return {
            "status": "success" if total_errors == 0 else "errors_found",
            "total_errors": total_errors,
            "total_formulas": total_formulas,
            "error_summary": summary,
        }
    except Exception as error:
        return {"error": str(error)}


def main():
    if len(sys.argv) < 2:
        print("Usage: python recalc.py <excel_file> [timeout_seconds]")
        raise SystemExit(1)
    timeout = int(sys.argv[2]) if len(sys.argv) > 2 else 30
    print(json.dumps(recalc(sys.argv[1], timeout), indent=2))


if __name__ == "__main__":
    main()
